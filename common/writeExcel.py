import os,sys
import shutil,time
from openpyxl.styles import Font
from openpyxl.styles import colors
import configparser as cparser
from openpyxl import load_workbook

from config import setting
sys.path.append(os.path.dirname(os.path.dirname(__file__)))

#读取config.ini配置
cf = cparser.ConfigParser()
cf.read(setting.TEST_CONFIG,encoding="utf-8")

class writeExcel():
    """写入excel"""
    def __init__(self,fileName):
        #需要把文件后缀上测试结果时间
        #获取当前时间
        now = time.strftime("%Y-%m-%d-%H-%M")
        splitFilename = fileName.split(".")
        self.filename =splitFilename[0]+"_"+now+"."+splitFilename[1]
        print(self.filename)
        if not os.path.exists(self.filename):
            shutil.copyfile(setting.SOURCE_FILE,setting.TARGET_FILE)
        #加载文件
        self.wb = load_workbook(self.filename)
        #获取当前工作表
        self.ws = self.wb.active

    def write_data(self,row_n,value):
        """
        excel中写入测试结果
        @param row_n: 数据所在行
        @param value: 测试结果值
        @return: 空
        """
        #设置格式
        font_green = Font(color=colors.COLOR_INDEX[3],bold=True)
        font_red = Font(color=colors.COLOR_INDEX[2],bold=True)
        # 获取所在行数
        L_n = "L"+ str(row_n)
        if value == "PASS":
            # 写入数据
            self.ws.cell(row_n,13,value)
            self.ws[L_n].font = font_green
        if value =="FAIL":
            self.ws.cell(row_n,13,value)
            self.ws[L_n].font = font_red
        self.wb.save(self.filename)